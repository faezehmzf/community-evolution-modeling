"""Synthetic schema-compatible xlsx fixtures for adapter tests (no real data)."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook

from tdmec_diagnostics.schema_contracts import (
    DATASET_A_DOCUMENTED_COLUMNS,
    DATASET_A_SHEET_NAME,
    DATASET_B_COLUMN_ORDER,
    DATASET_B_SHEET_NAME,
)


def write_dataset_b_fixture(path: Path, rows: List[Dict]) -> Path:
    """Write a Dataset-B-schema workbook (Sheet1, 10 columns)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = DATASET_B_SHEET_NAME
    ws.append(list(DATASET_B_COLUMN_ORDER))
    for r in rows:
        ws.append([r.get(c) for c in DATASET_B_COLUMN_ORDER])
    wb.save(path)
    wb.close()
    return path


def write_dataset_a_fixture(path: Path, rows: List[Dict]) -> Path:
    """Write a Dataset-A-schema workbook (tweets sheet, documented columns)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = DATASET_A_SHEET_NAME
    ws.append(list(DATASET_A_DOCUMENTED_COLUMNS))
    for r in rows:
        ws.append([r.get(c) for c in DATASET_A_DOCUMENTED_COLUMNS])
    wb.save(path)
    wb.close()
    return path


def minimal_dataset_b_rows() -> List[Dict]:
    return [
        {
            "id": "8000000000000000001",
            "created_at": "1519862400",  # 2018-03-01 UTC approx
            "user": "{'id': 1001, 'followers': 1, 'username': 'u1', 'title': None, 'political_category': None}",
            "text": "hello world node text",
            "likes": 0,
            "retweets": 0,
            "reply_count": 0,
            "quoted_count": 0,
            "bookmarks": 0,
            "views": 0,
        },
        {
            "id": "8000000000000000001",  # concordant duplicate
            "created_at": "1519862400",
            "user": "{'id': 1001, 'followers': 1, 'username': 'u1', 'title': None, 'political_category': None}",
            "text": "hello world node text",
            "likes": 0,
            "retweets": 0,
            "reply_count": 0,
            "quoted_count": 0,
            "bookmarks": 0,
            "views": 0,
        },
        {
            "id": "8000000000000000002",
            "created_at": "1519862400",
            "user": "{'id': 999999, 'followers': 1, 'username': 'out', 'title': None, 'political_category': None}",
            "text": "outside universe",
            "likes": 0,
            "retweets": 0,
            "reply_count": 0,
            "quoted_count": 0,
            "bookmarks": 0,
            "views": 0,
        },
        {
            "id": "",
            "created_at": "1519862400",
            "user": "{'id': 1002, 'followers': 1, 'username': 'u2', 'title': None, 'political_category': None}",
            "text": "",
            "likes": 0,
            "retweets": 0,
            "reply_count": 0,
            "quoted_count": 0,
            "bookmarks": 0,
            "views": 0,
        },
    ]


def minimal_dataset_a_rows() -> List[Dict]:
    base = {c: None for c in DATASET_A_DOCUMENTED_COLUMNS}
    rows = []
    r1 = dict(base)
    r1.update(
        {
            "id": 1.548256971999941e18,  # float-lossy untrusted
            "created_at": 1514764800,  # 2018-01-01
            "user": "{'id': 1001, 'username': 'a1'}",
            "text": "mention edge event",
            "user_mentions": "[{'id': 1002}]",
            "retweeted_status": None,
            "reply_status": None,
            "quoted_status": None,
        }
    )
    rows.append(r1)
    r2 = dict(base)
    r2.update(
        {
            "id": 1.548256971999942e18,
            "created_at": 1514764800,
            "user": "{'id': 1001, 'username': 'a1'}",
            "text": "self loop",
            "user_mentions": "[{'id': 1001}]",
        }
    )
    rows.append(r2)
    r3 = dict(base)
    r3.update(
        {
            "id": 1.548256971999943e18,
            "created_at": 1522540800,  # 2018-04-01
            "user": "{'id': 1002, 'username': 'a2'}",
            "text": None,
            "retweeted_status": "{'user': {'id': 1001}}",
        }
    )
    rows.append(r3)
    return rows


def write_unsupported_schema_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["foo", "bar"])
    ws.append([1, 2])
    wb.save(path)
    wb.close()
    return path
