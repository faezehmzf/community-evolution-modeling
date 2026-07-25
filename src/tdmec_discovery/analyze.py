"""Deep, single-file forensic analysis of a tweet/status workbook.

Uses the fast ``python-calamine`` reader when available (falls back to openpyxl)
and processes ONE workbook at a time. Designed for the ~1M-row, 100-270 MB
Dataset A / Dataset B files.

Nothing here loads more than one workbook into memory at once, and no raw text
is returned - only counts, hashes, aggregate statistics, and redacted samples.
"""
from __future__ import annotations

import datetime as _dt
import re
from typing import Dict, List, Optional, Sequence

from .fields import classify_columns

# Extract the leading numeric id from a ``user`` blob string:
#   "{'id': 1662724992054829056, 'followers': 235, 'username': 'x', ...}"
_USER_ID_RE = re.compile(r"'id'\s*:\s*(\d+)")
_USERNAME_RE = re.compile(r"'username'\s*:\s*'([^']*)'")


def _read_sheet(path: str):
    """Return (sheet_name, header, row_iter_factory) using the fastest engine."""
    try:
        from python_calamine import CalamineWorkbook

        wb = CalamineWorkbook.from_path(path)
        name = wb.sheet_names[0]
        ws = wb.get_sheet_by_index(0)
        data = ws.to_python(skip_empty_area=True)
        header = [str(h) if h is not None else f"__col{i}" for i, h in enumerate(data[0])]
        return name, header, data[1:]
    except Exception:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header_raw = next(it)
        header = [str(h) if h is not None else f"__col{i}" for i, h in enumerate(header_raw)]
        return ws.title, header, list(it)


def _parse_epoch(v) -> Optional[_dt.datetime]:
    if v is None or v == "":
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if x > 1e12:
        x /= 1000.0
    if 1_000_000_000 < x < 4_102_444_800:  # ~2001..2100
        try:
            return _dt.datetime.utcfromtimestamp(x)
        except (OverflowError, OSError, ValueError):
            return None
    return None


def _guess_dtype(values: Sequence) -> str:
    seen = set()
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            seen.add("bool")
        elif isinstance(v, int):
            seen.add("int")
        elif isinstance(v, float):
            seen.add("float")
        elif isinstance(v, (_dt.datetime, _dt.date)):
            seen.add("datetime")
        elif isinstance(v, str):
            s = v.strip()
            if s.startswith("{") or s.startswith("["):
                seen.add("json_blob")
            elif re.fullmatch(r"-?\d+", s):
                seen.add("int_str")
            else:
                seen.add("str")
        else:
            seen.add(type(v).__name__)
        if len(seen) > 3:
            break
    if not seen:
        return "empty"
    if seen <= {"int", "float"}:
        return "numeric"
    return "|".join(sorted(seen))


def deep_inspect(path: str, extract_edges: bool = False,
                 sample_rows: int = 3) -> dict:
    """Full single-pass forensic stats for one workbook."""
    sheet, header, rows = _read_sheet(path)
    ncols = len(header)
    idx = {c: i for i, c in enumerate(header)}
    n = len(rows)

    roles = classify_columns(header)

    def col(name: str) -> Optional[int]:
        return idx.get(name)

    null_counts = {c: 0 for c in header}
    empty_text = 0
    # dtype sample: first 2000 rows
    dtype_samples = {c: [] for c in header}

    id_col = "id" if "id" in idx else None
    ts_col = "created_at" if "created_at" in idx else (
        "timestamp" if "timestamp" in idx else None)
    user_col = "user" if "user" in idx else None
    text_col = "text" if "text" in idx else None

    tweet_ids_seen = set()
    tweet_id_dupes = 0
    author_ids = set()
    usernames = set()
    ts_min = None
    ts_max = None
    ts_invalid = 0

    # edge target sets (Dataset A only)
    edge_cols = {
        "reply": "reply_status",
        "retweet": "retweeted_status",
        "quote": "quoted_status",
        "mention": "user_mentions",
    }
    edge_nonnull = {k: 0 for k in edge_cols}

    for ri, row in enumerate(rows):
        rowlen = len(row)
        for c, i in idx.items():
            v = row[i] if i < rowlen else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                null_counts[c] += 1
            elif ri < 2000:
                dtype_samples[c].append(v)
        if text_col is not None:
            tv = row[idx[text_col]] if idx[text_col] < rowlen else None
            if tv is None or (isinstance(tv, str) and tv.strip() == ""):
                empty_text += 1
        if id_col is not None:
            v = row[idx[id_col]] if idx[id_col] < rowlen else None
            if v is not None and v != "":
                key = str(v)
                if key in tweet_ids_seen:
                    tweet_id_dupes += 1
                else:
                    tweet_ids_seen.add(key)
        if user_col is not None:
            v = row[idx[user_col]] if idx[user_col] < rowlen else None
            if isinstance(v, str):
                m = _USER_ID_RE.search(v)
                if m:
                    author_ids.add(m.group(1))
                mu = _USERNAME_RE.search(v)
                if mu and mu.group(1):
                    usernames.add(mu.group(1))
        if ts_col is not None:
            v = row[idx[ts_col]] if idx[ts_col] < rowlen else None
            dtv = _parse_epoch(v)
            if dtv is None:
                if v is not None and v != "":
                    ts_invalid += 1
            else:
                if ts_min is None or dtv < ts_min:
                    ts_min = dtv
                if ts_max is None or dtv > ts_max:
                    ts_max = dtv
        if extract_edges:
            for role, cname in edge_cols.items():
                ci = idx.get(cname)
                if ci is None:
                    continue
                v = row[ci] if ci < rowlen else None
                if isinstance(v, str) and _USER_ID_RE.search(v):
                    edge_nonnull[role] += 1

    column_dtypes = {c: _guess_dtype(dtype_samples[c]) for c in header}

    result = {
        "sheet": sheet,
        "n_data_rows": n,
        "n_columns": ncols,
        "columns": header,
        "column_dtypes": column_dtypes,
        "null_counts": null_counts,
        "candidate_fields": {k: v for k, v in roles.items() if v},
        "tweet_id_field": id_col,
        "tweet_id_distinct": len(tweet_ids_seen) if id_col else None,
        "tweet_id_duplicate_rows": tweet_id_dupes if id_col else None,
        "timestamp_field": ts_col,
        "ts_min": ts_min.isoformat() if ts_min else None,
        "ts_max": ts_max.isoformat() if ts_max else None,
        "ts_invalid": ts_invalid if ts_col else None,
        "author_id_field": f"{user_col}.id" if user_col else None,
        "approx_unique_author_ids": len(author_ids) if user_col else None,
        "approx_unique_usernames": len(usernames) if user_col else None,
        "empty_text_rows": empty_text if text_col else None,
    }
    if extract_edges:
        result["edge_nonnull_counts"] = edge_nonnull
    # return author id set separately for corpus-level reconciliation (not serialized here)
    result["_author_ids"] = author_ids
    return result
