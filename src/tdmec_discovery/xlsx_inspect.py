"""Excel workbook inspection utilities.

All reads are streaming and read-only. Workbooks are opened with
``read_only=True`` so a single file is processed without loading the whole
workbook into memory. Callers are expected to inspect ONE file at a time and
delete the local copy afterwards.
"""
from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence


def workbook_sheet_names(path: str | Path) -> List[str]:
    """Return sheet names without reading any cell data."""
    import openpyxl

    wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _infer_type(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (_dt.datetime, _dt.date)):
        return "datetime"
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return "empty_str"
        return "str"
    return type(value).__name__


@dataclass
class ColumnStats:
    name: str
    order: int
    type_counts: Dict[str, int] = field(default_factory=dict)
    null_count: int = 0
    non_null_count: int = 0

    def candidate_dtype(self) -> str:
        counts = {k: v for k, v in self.type_counts.items() if k not in ("null", "empty_str")}
        if not counts:
            return "empty"
        # Collapse int/float -> numeric when mixed.
        top = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)
        if {"int", "float"} >= set(counts) and len(counts) > 1:
            return "numeric"
        return top[0][0]


@dataclass
class SheetInspection:
    sheet: str
    n_data_rows: int
    columns: List[str]
    column_stats: Dict[str, ColumnStats]
    sample_rows: List[dict]
    key_stats: Dict[str, dict] = field(default_factory=dict)
    error: Optional[str] = None

    def to_summary(self, include_samples: bool = False) -> dict:
        out = {
            "sheet": self.sheet,
            "n_data_rows": self.n_data_rows,
            "n_columns": len(self.columns),
            "columns": self.columns,
            "column_types": {c: self.column_stats[c].candidate_dtype() for c in self.columns},
            "null_counts": {c: self.column_stats[c].null_count for c in self.columns},
            "key_stats": self.key_stats,
            "error": self.error,
        }
        if include_samples:
            out["sample_rows"] = self.sample_rows
        return out


def inspect_sheet(
    path: str | Path,
    sheet: Optional[str] = None,
    sample_rows: int = 20,
    key_columns: Optional[Sequence[str]] = None,
    dtype_scan_rows: Optional[int] = None,
    collect_distinct: Optional[Sequence[str]] = None,
    timestamp_columns: Optional[Sequence[str]] = None,
) -> SheetInspection:
    """Single streaming pass over one sheet.

    Parameters
    ----------
    sample_rows: number of leading data rows to retain as (redactable) samples.
    key_columns: columns to track for exact duplicate counts (e.g. tweet id).
    dtype_scan_rows: if set, only scan this many rows for dtype/null inference
        (row count still reflects the full pass). If None, scan every row.
    collect_distinct: columns for which to track approximate distinct counts.
    timestamp_columns: columns for which to track min/max + invalid counts.
    """
    import openpyxl

    key_columns = list(key_columns or [])
    collect_distinct = list(collect_distinct or [])
    timestamp_columns = list(timestamp_columns or [])

    wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return SheetInspection(sheet=ws.title, n_data_rows=0, columns=[], column_stats={}, sample_rows=[])

        columns = [str(h) if h is not None else f"__col{i}" for i, h in enumerate(header)]
        stats = {c: ColumnStats(name=c, order=i) for i, c in enumerate(columns)}
        col_index = {c: i for i, c in enumerate(columns)}

        key_seen: Dict[str, set] = {c: set() for c in key_columns if c in col_index}
        key_dups: Dict[str, int] = {c: 0 for c in key_seen}
        distinct: Dict[str, set] = {c: set() for c in collect_distinct if c in col_index}
        ts_min: Dict[str, object] = {}
        ts_max: Dict[str, object] = {}
        ts_invalid: Dict[str, int] = {c: 0 for c in timestamp_columns if c in col_index}
        empty_text_cols: Dict[str, int] = {}

        samples: List[dict] = []
        n = 0
        for row in rows_iter:
            n += 1
            do_types = dtype_scan_rows is None or n <= dtype_scan_rows
            for c, i in col_index.items():
                val = row[i] if i < len(row) else None
                if do_types:
                    st = stats[c]
                    t = _infer_type(val)
                    st.type_counts[t] = st.type_counts.get(t, 0) + 1
                    if t in ("null", "empty_str"):
                        st.null_count += 1
                    else:
                        st.non_null_count += 1
            for c in key_seen:
                v = row[col_index[c]] if col_index[c] < len(row) else None
                if v is not None:
                    if v in key_seen[c]:
                        key_dups[c] += 1
                    else:
                        key_seen[c].add(v)
            for c in distinct:
                v = row[col_index[c]] if col_index[c] < len(row) else None
                if v is not None and v != "":
                    distinct[c].add(v)
            for c in ts_invalid:
                v = row[col_index[c]] if col_index[c] < len(row) else None
                parsed = _coerce_ts(v)
                if parsed is None:
                    if v is not None and v != "":
                        ts_invalid[c] += 1
                else:
                    if c not in ts_min or parsed < ts_min[c]:
                        ts_min[c] = parsed
                    if c not in ts_max or parsed > ts_max[c]:
                        ts_max[c] = parsed
            if len(samples) < sample_rows:
                samples.append({c: _jsonable(row[col_index[c]] if col_index[c] < len(row) else None) for c in columns})

        key_stats: Dict[str, dict] = {}
        for c in key_seen:
            key_stats[c] = {"distinct": len(key_seen[c]), "duplicate_rows": key_dups[c]}
        for c in distinct:
            key_stats.setdefault(c, {})["approx_distinct"] = len(distinct[c])
        for c in ts_invalid:
            key_stats.setdefault(c, {}).update(
                {
                    "ts_min": _jsonable(ts_min.get(c)),
                    "ts_max": _jsonable(ts_max.get(c)),
                    "ts_invalid": ts_invalid[c],
                }
            )

        return SheetInspection(
            sheet=ws.title,
            n_data_rows=n,
            columns=columns,
            column_stats=stats,
            sample_rows=samples,
            key_stats=key_stats,
        )
    finally:
        wb.close()


def _coerce_ts(v):
    if v is None or v == "":
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v if isinstance(v, _dt.datetime) else _dt.datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)):
        # Heuristic: treat large numbers as unix epoch seconds / ms.
        try:
            x = float(v)
            if x > 1e12:  # milliseconds
                x /= 1000.0
            if 0 < x < 4102444800:  # up to year 2100
                return _dt.datetime.utcfromtimestamp(x)
        except Exception:
            return None
        return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d",
            "%a %b %d %H:%M:%S %z %Y",  # Twitter created_at
            "%Y-%m-%d %H:%M:%S%z",
        ):
            try:
                return _dt.datetime.strptime(s, fmt).replace(tzinfo=None)
            except (ValueError, TypeError):
                continue
    return None


def _jsonable(v):
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v
