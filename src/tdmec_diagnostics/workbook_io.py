"""Streaming workbook helpers (bounded memory; source files never modified)."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterator, List, Sequence, Tuple


class UnsupportedSchemaError(ValueError):
    """Raised when a workbook schema does not match the documented contract."""


def validate_required_columns(
    header: Sequence[str],
    required: Sequence[str],
    *,
    adapter_id: str,
    allow_extra: bool = True,
) -> Dict[str, int]:
    """Return column name → index. Reject missing required columns (no guessing)."""
    cols = [str(h) if h is not None else "" for h in header]
    index = {name: i for i, name in enumerate(cols)}
    missing = [c for c in required if c not in index]
    if missing:
        raise UnsupportedSchemaError(
            f"adapter={adapter_id} missing required columns: {missing}; "
            f"observed_columns={cols}"
        )
    if not allow_extra:
        extra = [c for c in cols if c and c not in required]
        if extra:
            raise UnsupportedSchemaError(
                f"adapter={adapter_id} unexpected columns: {extra}"
            )
    return {c: index[c] for c in required}


def iter_xlsx_rows(
    path: str | Path,
    *,
    expected_sheet: str | None = None,
    start_after_row_number: int = 1,
) -> Tuple[str, List[str], Iterator[Tuple]]:
    """Stream workbook rows without modifying the source file.

    Prefers openpyxl read_only iteration (true streaming). Falls back to
    python-calamine which materializes one sheet (documented trade-off).
    """
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"workbook not found: {path.name}")

    # Prefer streaming read_only openpyxl
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            if expected_sheet is not None:
                if expected_sheet not in wb.sheetnames:
                    raise UnsupportedSchemaError(
                        f"expected sheet {expected_sheet!r} not in {wb.sheetnames}"
                    )
                ws = wb[expected_sheet]
                sheet_name = expected_sheet
            else:
                sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]
            header_rows = ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
            header_raw = next(header_rows)
            header = [
                str(h) if h is not None else f"__col{i}"
                for i, h in enumerate(header_raw)
            ]
            first_data_row = max(2, int(start_after_row_number) + 1)
            rows = ws.iter_rows(
                min_row=first_data_row,
                values_only=True,
            )

            def _gen() -> Iterator[Tuple]:
                try:
                    for row in rows:
                        yield tuple(row)
                finally:
                    wb.close()

            return sheet_name, header, _gen()
        except Exception:
            wb.close()
            raise
    except UnsupportedSchemaError:
        raise
    except Exception:
        pass

    # Fallback: calamine (one-sheet materialization)
    from python_calamine import CalamineWorkbook

    wb2 = CalamineWorkbook.from_path(str(path))
    names = list(wb2.sheet_names)
    if expected_sheet is not None:
        if expected_sheet not in names:
            raise UnsupportedSchemaError(
                f"expected sheet {expected_sheet!r} not in {names}"
            )
        idx = names.index(expected_sheet)
        sheet_name = expected_sheet
    else:
        idx = 0
        sheet_name = names[0]
    data = wb2.get_sheet_by_index(idx).to_python(skip_empty_area=True)
    if not data:
        raise UnsupportedSchemaError(f"empty workbook sheet in {path.name}")
    header = [str(h) if h is not None else f"__col{i}" for i, h in enumerate(data[0])]

    first_data_index = max(1, int(start_after_row_number))

    def _gen2() -> Iterator[Tuple]:
        for row in data[first_data_index:]:
            yield tuple(row)

    return sheet_name, header, _gen2()
